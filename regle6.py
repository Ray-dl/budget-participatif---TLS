import openpyxl

table = openpyxl.load_workbook('x:/xxxx/regle 3 et 6.xlsx')# ouvrire file, "r" read
sheetname = table.sheetnames
sheet = table[sheetname[0]] #position de sheet dans excel
nbrow = sheet.max_row  #count nombre de lignes
nbcol = sheet.max_column  #count nombre de colones

def proportionnel():
    sheet.cell(1, 9, 'proportionnel')  # inserer titre 'proportionnel' a cellule(1,9)/I1
    for i in range(2, nbrow+1):
         satisfaction= sheet.cell(i,7).value*sheet.cell(i,8).value # satisfaction = nombre de votant * cout de projet choisi
         sheet.cell(i, 9).value = satisfaction / sheet.cell(i, 8).value # proportionnel = satisfaction / cout, cout dans colone 8
         table.save('regle 3 et 6.xlsx')
         #print(sheet.cell(i, 9).value)





def regle6_cout_proportionnel():
    # 1. fonction transformer valeur a liste et dictionnaire
    liste_proportionnel = []
    liste_idee = []
    liste_cout = []
    dict1={}
    for i in range(2, nbrow + 1):
        proportionnel = sheet.cell(i, 9).value
        idee = sheet.cell(i, 1).value
        cout = int(sheet.cell(i, 8).value)
        liste_idee.append(idee)   # ajouter les valeurs à une liste
        liste_proportionnel.append(proportionnel)
        liste_cout.append(cout)
        dictionary1 =dict(zip(liste_proportionnel,liste_cout))  #mettre deux liste dans un dictionnaire, key est proportionnel, value est cout
        dictionary2 = dict(zip(liste_proportionnel, liste_idee)) #mettre deux liste dans un dictionnaire, key est proportionnel, value est le nom de projet
        dict1.update(dictionary1) #mettre le dictionary1 à un dictionnaire libre'dict1'
    dict1=sorted(dict1.keys(),reverse=True) #classer le dict1 par ordre décroissant de key dans le dictionnaire, en outre, le type de dictionaire devient liste
    #print(dict1)


    # 2  calculer cout
    somme_cout = 0
    for i in range(0, len(dict1)):
            prop=dict1[i]   #donner l'element de liste dict1 à prop
            if somme_cout + dictionary1[prop] <= 1000000:  #verifier la somme de cout est inferieur ou egale a 1000000
                somme_cout += dictionary1[prop]   #si inferieur ou egale a 1000000, plus la valeur de cout dans somme——cout
                print(dictionary2[prop],"Cout:",dictionary1[prop],"proportionnel:", prop,"somme-cout:",somme_cout)
            else:
                continue




print("Idee chosir sont: ")
# call fonction:
proportionnel()
regle6_cout_proportionnel()

